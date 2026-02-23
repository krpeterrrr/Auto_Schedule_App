# main.py
# 실행: python -m streamlit run main.py
# 필요 패키지: pip install streamlit openpyxl

import streamlit as st
from collections import defaultdict
import csv
import io
import html

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

# =========================
# 설정
# =========================
DAYS = ["월", "화", "수", "목", "금"]

def generate_slots(days, start_h, end_h):
    return [f"{d}_{h}" for d in days for h in range(start_h, end_h)]

def slot_to_display(slot):
    d, h = slot.split("_")
    h = int(h)
    return d, f"{h:02d}:00-{h+1:02d}:00"

def consecutive_run_lengths(bool_list):
    runs, cur = [], 0
    for v in bool_list:
        if v:
            cur += 1
        else:
            if cur:
                runs.append(cur)
            cur = 0
    if cur:
        runs.append(cur)
    return runs

def max_consecutive_if_assigned(person_slots_set, day, start_h, end_h):
    arr = [(f"{day}_{h}") in person_slots_set for h in range(start_h, end_h)]
    runs = consecutive_run_lengths(arr)
    return max(runs) if runs else 0

# =========================
# HTML 테이블 렌더러 (pandas 없이)
# =========================
def render_table(rows, columns, title=None):
    if title:
        st.markdown(f"**{title}**")

    st.markdown(
        """
        <style>
        .tbl-wrap { width: 100%; overflow-x: auto; }
        table.tbl {
            border-collapse: collapse;
            width: 100%;
            min-width: 520px;
            font-size: 14px;
        }
        .tbl th, .tbl td {
            border: 1px solid #E5E7EB;
            padding: 8px 10px;
            text-align: center;
            vertical-align: middle;
            white-space: nowrap;
        }
        .tbl th {
            background: #F9FAFB;
            font-weight: 700;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    thead = "<tr>" + "".join(f"<th>{html.escape(col)}</th>" for col in columns) + "</tr>"
    body_rows = []
    for r in rows:
        tds = []
        for col in columns:
            val = r.get(col, "")
            tds.append(f"<td>{html.escape(str(val))}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")

    table_html = f"""
    <div class="tbl-wrap">
      <table class="tbl">
        <thead>{thead}</thead>
        <tbody>{''.join(body_rows)}</tbody>
      </table>
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)

# =========================
# 자동 배정(백트래킹 + 휴리스틱)
# =========================
def auto_assign(people, slots, availability, start_h, end_h, max_consec=2, min_each=1):
    candidates = {s: [p for p in people if availability[p].get(s, False)] for s in slots}

    zero_slots = [s for s in slots if len(candidates[s]) == 0]
    if zero_slots:
        return None, {"reason": "후보가 0명인 시간대가 있어 배정이 불가능합니다.", "zero_slots": zero_slots}

    ordered_slots = sorted(slots, key=lambda s: (len(candidates[s]), s))

    assigned = {}
    person_slots = {p: set() for p in people}
    person_total = {p: 0 for p in people}

    def can_assign(p, s):
        d, _ = s.split("_")
        tmp = set(person_slots[p])
        tmp.add(s)
        return max_consecutive_if_assigned(tmp, d, start_h, end_h) <= max_consec

    def person_priority(p):
        # 공평: 현재 근무 적은 사람 우선, 동점이면 입력 순서
        return (person_total[p], people.index(p))

    def backtrack(i):
        if i == len(ordered_slots):
            return all(person_total[p] >= min_each for p in people)

        s = ordered_slots[i]
        cand = sorted(candidates[s], key=person_priority)

        for p in cand:
            if not can_assign(p, s):
                continue

            assigned[s] = p
            person_slots[p].add(s)
            person_total[p] += 1

            if backtrack(i + 1):
                return True

            person_total[p] -= 1
            person_slots[p].remove(s)
            del assigned[s]

        return False

    if not backtrack(0):
        return None, {
            "reason": "제약(연속 근무 제한 / 최소 근무시간)을 만족하며 전체 시간대를 커버하는 배정을 찾지 못했습니다.",
            "hint": "가능 시간을 더 체크하거나(특히 가능 인원이 적은 시간대), 인원을 늘리거나, 연속 제한을 완화해 보세요."
        }

    return assigned, None

# =========================
# (신규) 엑셀 생성(.xlsx)
# =========================
def build_assignment_excel(assigned, people_in_order, start_h, end_h):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "배정결과"

    headers = ["시간"] + DAYS
    ws.append(headers)

    # 데이터
    for h in range(start_h, end_h):
        row = [f"{h:02d}:00-{h+1:02d}:00"]
        for d in DAYS:
            row.append(assigned.get(f"{d}_{h}", ""))
        ws.append(row)

    # 스타일
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(bold=True)

    # 헤더 스타일
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 본문 스타일
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

    # 열 너비
    ws.column_dimensions["A"].width = 14
    for idx in range(2, 2 + len(DAYS)):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    # 2번째 시트: 개인별 총 근무시간(입력 순서)
    ws2 = wb.create_sheet("개인별근무시간")
    ws2.append(["이름", "총 근무시간"])

    totals = defaultdict(int)
    for _, p in assigned.items():
        totals[p] += 1

    for p in people_in_order:
        ws2.append([p, int(totals.get(p, 0))])

    # 스타일 적용
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r in range(2, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    # 메모리로 저장
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()

# =========================
# Streamlit 초기화
# =========================
st.set_page_config(page_title="시간표 배정", layout="wide")
st.title("시간표 배정 (가능시간 취합 + 전체 커버 자동 배정)")

if "start_h" not in st.session_state:
    st.session_state.start_h = 10
if "end_h" not in st.session_state:
    st.session_state.end_h = 17
if "people" not in st.session_state:
    st.session_state.people = []
if "availability" not in st.session_state:
    st.session_state.availability = {}
if "assigned" not in st.session_state:
    st.session_state.assigned = None

# =========================
# 0) 시간 범위 설정
# =========================
st.subheader("0) 시간 범위 설정 (전체 커버 대상)")
c1, c2, c3 = st.columns([1, 1, 2])
with c1:
    start_h = st.number_input("시작 시간(시)", 0, 23, int(st.session_state.start_h), 1)
with c2:
    end_h = st.number_input("끝 시간(시) (포함 X)", 1, 24, int(st.session_state.end_h), 1)
with c3:
    st.caption("예) 10~17 → 10-11 ... 16-17 (7칸). 월~금 전체를 모두 커버합니다.")

if end_h <= start_h:
    st.error("끝 시간은 시작 시간보다 커야 합니다.")
    st.stop()

start_h, end_h = int(start_h), int(end_h)
new_slots = generate_slots(DAYS, start_h, end_h)

# 시간범위 변경 시 availability 동기화
if (start_h != st.session_state.start_h) or (end_h != st.session_state.end_h):
    for p in st.session_state.people:
        old = st.session_state.availability.get(p, {})
        st.session_state.availability[p] = {s: bool(old.get(s, False)) for s in new_slots}
    st.session_state.start_h = start_h
    st.session_state.end_h = end_h
    st.session_state.assigned = None

SLOTS = new_slots

# =========================
# 1) 사람 관리 (Enter로 추가)
# =========================
st.divider()
st.subheader("1) 사람 관리")

def add_person_from_input():
    name = (st.session_state.get("new_person_name", "") or "").strip()
    if not name:
        return
    if name in st.session_state.people:
        st.session_state.new_person_name = ""
        return
    st.session_state.people.append(name)
    st.session_state.availability[name] = {s: False for s in SLOTS}
    st.session_state.new_person_name = ""
    st.session_state.assigned = None

left, right = st.columns([2, 1])
with left:
    st.text_input(
        "이름 입력 후 Enter",
        key="new_person_name",
        placeholder="예: 김공명",
        on_change=add_person_from_input,
    )

with right:
    if st.session_state.people:
        del_target = st.selectbox("삭제할 사람", st.session_state.people, key="del_target")
        if st.button("선택 삭제", use_container_width=True):
            st.session_state.people.remove(del_target)
            st.session_state.availability.pop(del_target, None)
            st.session_state.assigned = None

if not st.session_state.people:
    st.info("사람을 추가한 뒤, 아래에서 선택하여 가능 시간을 체크하세요.")
    st.stop()

# =========================
# 2) 가능 시간 입력 (사람 리스트 박스 스크롤 + 검색)
# =========================
st.divider()
st.subheader("2) 가능 시간 입력")

col_people, col_table = st.columns([1, 4], vertical_alignment="top")

with col_people:
    st.markdown("#### 사람 목록")

    q = st.text_input("이름 검색", key="people_search", placeholder="검색어 입력")
    filtered = st.session_state.people
    if q.strip():
        qq = q.strip().lower()
        filtered = [p for p in st.session_state.people if qq in p.lower()]

    if not filtered:
        st.warning("검색 결과가 없습니다. 검색어를 지우면 전체가 다시 표시됩니다.")
        st.stop()

    box = st.container(height=520, border=True)

    prev = st.session_state.get("selected_person_radio", None)
    idx = filtered.index(prev) if prev in filtered else 0

    with box:
        selected_person = st.radio(
            " ",
            options=filtered,
            index=idx,
            label_visibility="collapsed",
            key="selected_person_radio"
        )

with col_table:
    st.markdown(f"#### {selected_person} 가능 시간 선택")

    day_cols = st.columns(len(DAYS))
    for i, day in enumerate(DAYS):
        with day_cols[i]:
            st.markdown(f"**{day}**")
            for h in range(start_h, end_h):
                slot = f"{day}_{h}"
                key = f"cb_{selected_person}_{day}_{h}"
                val = bool(st.session_state.availability[selected_person].get(slot, False))
                checked = st.checkbox(
                    f"{h:02d}:00-{h+1:02d}:00",
                    value=val,
                    key=key
                )
                st.session_state.availability[selected_person][slot] = bool(checked)

    st.caption("입력(체크)은 제한 없음. 연속 2시간 제한/최소 1시간은 '자동 배정'에서만 적용됩니다.")

# =========================
# 3) 집계 (요일별) — HTML 테이블
# =========================
st.divider()
st.subheader("3) 시간대별 가능 인원 집계 (요일별)")

pick_day = st.selectbox("요일 선택", DAYS, key="count_day")
count_rows = []
for h in range(start_h, end_h):
    slot = f"{pick_day}_{h}"
    cnt = sum(st.session_state.availability[p].get(slot, False) for p in st.session_state.people)
    _, disp = slot_to_display(slot)
    count_rows.append({"시간": disp, "가능 인원": cnt})

render_table(count_rows, columns=["시간", "가능 인원"])

# =========================
# 4) 자동 배정
# =========================
st.divider()
st.subheader("4) 자동 근무 배정 (전체 시간대 커버)")

c1, c2, c3 = st.columns([1, 1, 2])
with c1:
    min_each = st.number_input("사람당 최소 근무시간(시간)", 0, 200, 1, 1)
with c2:
    max_consec = st.number_input("하루 최대 연속 근무시간(시간)", 1, 12, 2, 1)
with c3:
    st.caption("모든 슬롯(월~금 × 시간칸)에 1명씩 배정합니다.")

if st.button("자동 배정 실행", type="primary", use_container_width=True):
    assigned, err = auto_assign(
        people=list(st.session_state.people),
        slots=SLOTS,
        availability=st.session_state.availability,
        start_h=start_h,
        end_h=end_h,
        max_consec=int(max_consec),
        min_each=int(min_each),
    )
    if err:
        st.session_state.assigned = None
        st.error(err["reason"])
        if "zero_slots" in err:
            st.write("후보 0명 시간대(아무도 체크 안 함):")
            for s in err["zero_slots"]:
                d, disp = slot_to_display(s)
                st.write(f"- {d} {disp}")
        if "hint" in err:
            st.info(err["hint"])
    else:
        st.session_state.assigned = assigned
        st.success("배정 완료!")

# =========================
# 5) 최종 결과 + 엑셀 다운로드
# =========================
st.divider()
st.subheader("5) 최종 배정 결과 (표 형태)")

if not st.session_state.assigned:
    st.info("아직 배정 결과가 없습니다. 위에서 '자동 배정 실행'을 눌러주세요.")
else:
    assigned = st.session_state.assigned

    table_rows = []
    for h in range(start_h, end_h):
        row = {"시간": f"{h:02d}:00-{h+1:02d}:00"}
        for d in DAYS:
            row[d] = assigned.get(f"{d}_{h}", "")
        table_rows.append(row)

    render_table(table_rows, columns=["시간"] + DAYS)

    # ✅ 엑셀 다운로드 버튼
    xlsx_bytes = build_assignment_excel(
        assigned=assigned,
        people_in_order=list(st.session_state.people),
        start_h=start_h,
        end_h=end_h
    )
    st.download_button(
        label="배정 결과 엑셀(.xlsx) 다운로드",
        data=xlsx_bytes,
        file_name="assignment_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # 개인별 총 근무시간(입력 순서)
    st.subheader("개인별 총 근무시간 (입력 순서)")
    totals = defaultdict(int)
    for _, p in assigned.items():
        totals[p] += 1
    load_rows = [{"이름": p, "총 근무시간": int(totals.get(p, 0))} for p in st.session_state.people]
    render_table(load_rows, columns=["이름", "총 근무시간"])